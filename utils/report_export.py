"""
report_export.py
----------------
Build SharePoint archive files for weekly/monthly offboard reports.

Formats:
  - CSV  (stdlib) – flat file with a Section column, easy to append/query
  - XLSX (openpyxl) – multi-sheet workbook for Excel

Env: REPORT_ARCHIVE_FORMAT = csv | xlsx | both  (default: both)
"""

from __future__ import annotations

import csv
import io
import os
from typing import Any, Dict, List, Tuple


def get_archive_formats() -> List[str]:
    raw = os.environ.get("REPORT_ARCHIVE_FORMAT", "both").strip().lower()
    if raw in ("csv", "xlsx"):
        return [raw]
    return ["xlsx", "csv"]


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def build_report_csv(
    *,
    report_date: str,
    period_label: str,
    period_start: str,
    new_no_ef: List[Dict[str, Any]],
    new_with_ef: List[Dict[str, Any]],
    overdue_no_ef: List[Dict[str, Any]],
    summary: Dict[str, Any],
) -> bytes:
    """Single CSV with Section column for long-term data handling."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Section",
        "ReportDate",
        "PeriodLabel",
        "PeriodStart",
        "DisplayName",
        "UserEmail",
        "OffboardDate",
        "UsageLocation",
        "ManagerEmail",
        "ForwardingAddress",
        "DaysElapsed",
        "Metric",
        "MetricValue",
    ])

    for key, label in (
        ("alerts", "Alerts"),
        ("extensions", "Extensions"),
        ("deletions", "Deletions"),
        ("total_active", "ActiveTracked"),
    ):
        writer.writerow([
            "Summary", report_date, period_label, period_start,
            "", "", "", "", "", "", "",
            label, summary.get(key, summary.get(label.lower(), "")),
        ])
    writer.writerow([
        "Summary", report_date, period_label, period_start,
        "", "", "", "", "", "", "",
        "NewNO_EF_Count", len(new_no_ef),
    ])
    writer.writerow([
        "Summary", report_date, period_label, period_start,
        "", "", "", "", "", "", "",
        "NewEF_Count", len(new_with_ef),
    ])
    writer.writerow([
        "Summary", report_date, period_label, period_start,
        "", "", "", "", "", "", "",
        "OverdueNO_EF_Count", len(overdue_no_ef),
    ])

    for r in new_no_ef:
        writer.writerow([
            "New_NO_EF", report_date, period_label, period_start,
            _cell(r.get("displayName")), _cell(r.get("userEmail")),
            _cell(r.get("offboardDate")), _cell(r.get("usageLocation")),
            _cell(r.get("managerEmail")), "", "", "", "",
        ])

    for r in new_with_ef:
        writer.writerow([
            "New_EF", report_date, period_label, period_start,
            _cell(r.get("displayName")), _cell(r.get("userEmail")),
            _cell(r.get("offboardDate")), _cell(r.get("usageLocation")),
            _cell(r.get("managerEmail")), _cell(r.get("forwardingAddress")),
            "", "", "",
        ])

    for r in overdue_no_ef:
        writer.writerow([
            "Overdue_NO_EF", report_date, period_label, period_start,
            _cell(r.get("displayName")), _cell(r.get("userEmail")),
            _cell(r.get("offboardDate")), _cell(r.get("usageLocation")),
            _cell(r.get("managerEmail")), "",
            _cell(r.get("daysElapsed")), "", "",
        ])

    # UTF-8 BOM so Excel opens CSV cleanly
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def build_report_xlsx(
    *,
    report_date: str,
    period_label: str,
    period_start: str,
    new_no_ef: List[Dict[str, Any]],
    new_with_ef: List[Dict[str, Any]],
    overdue_no_ef: List[Dict[str, Any]],
    summary: Dict[str, Any],
) -> bytes:
    """Multi-sheet Excel workbook."""
    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    ws.append(["ReportDate", report_date])
    ws.append(["PeriodLabel", period_label])
    ws.append(["PeriodStart", period_start])
    ws.append(["New_NO_EF", len(new_no_ef)])
    ws.append(["New_EF", len(new_with_ef)])
    ws.append(["Overdue_NO_EF", len(overdue_no_ef)])
    ws.append(["Alerts", summary.get("alerts", summary.get("alerted", ""))])
    ws.append(["Extensions", summary.get("extensions", "")])
    ws.append(["Deletions", summary.get("deletions", summary.get("deleted", ""))])
    ws.append(["ActiveTracked", summary.get("total_active", "")])

    ws_no = wb.create_sheet("New_NO_EF")
    ws_no.append(["DisplayName", "UserEmail", "OffboardDate", "UsageLocation", "ManagerEmail"])
    for r in new_no_ef:
        ws_no.append([
            r.get("displayName", ""), r.get("userEmail", ""),
            r.get("offboardDate", ""), r.get("usageLocation", ""),
            r.get("managerEmail", ""),
        ])

    ws_ef = wb.create_sheet("New_EF")
    ws_ef.append([
        "DisplayName", "UserEmail", "OffboardDate", "ForwardingAddress",
        "ManagerEmail", "UsageLocation",
    ])
    for r in new_with_ef:
        ws_ef.append([
            r.get("displayName", ""), r.get("userEmail", ""),
            r.get("offboardDate", ""), r.get("forwardingAddress", ""),
            r.get("managerEmail", ""), r.get("usageLocation", ""),
        ])

    ws_od = wb.create_sheet("Overdue_NO_EF")
    ws_od.append([
        "DisplayName", "UserEmail", "OffboardDate", "UsageLocation",
        "ManagerEmail", "DaysElapsed",
    ])
    for r in overdue_no_ef:
        ws_od.append([
            r.get("displayName", ""), r.get("userEmail", ""),
            r.get("offboardDate", ""), r.get("usageLocation", ""),
            r.get("managerEmail", ""), r.get("daysElapsed", ""),
        ])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_report_archives(
    *,
    report_date: str,
    period_label: str,
    period_start: str,
    new_no_ef: List[Dict[str, Any]],
    new_with_ef: List[Dict[str, Any]],
    overdue_no_ef: List[Dict[str, Any]],
    summary: Dict[str, Any],
) -> List[Tuple[str, bytes, str]]:
    """
    Returns list of (filename, content_bytes, content_type).
    """
    safe_label = period_label.replace(" ", "_")
    base = f"EF_{safe_label}_Report_{report_date}"
    kwargs = dict(
        report_date=report_date,
        period_label=period_label,
        period_start=period_start,
        new_no_ef=new_no_ef,
        new_with_ef=new_with_ef,
        overdue_no_ef=overdue_no_ef,
        summary=summary,
    )
    files: List[Tuple[str, bytes, str]] = []
    for fmt in get_archive_formats():
        if fmt == "csv":
            files.append((
                f"{base}.csv",
                build_report_csv(**kwargs),
                "text/csv; charset=utf-8",
            ))
        elif fmt == "xlsx":
            files.append((
                f"{base}.xlsx",
                build_report_xlsx(**kwargs),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ))
    return files
